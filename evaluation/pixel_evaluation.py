"""
Version: 1.5

Summary: Compute the similarity of a pair of images

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

    python3 pixel_evaluation.py -i /input/ -o /output/
    
"""

import subprocess, os
import sys
import argparse
import numpy as np 
import pathlib
import os
import glob
import shutil
from pathlib import Path
from scipy.spatial import KDTree
import cv2
import imutils




import pandas as pd


#from image_similarity_measures.evaluate import evaluation

import openpyxl



# generate foloder to store the output results
def mkdir(path):
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #printpath + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        
        print("{} path exists!\n".format(path))
        return False

def similarity_evaluation(image_ori, image_pred):
    
    file_name = Path(image_ori).name
            
            
    # metrics = ["rmse", "ssim"]
    dict_val = evaluation(org_img_path = image_ori, pred_img_path = image_pred, metrics = ["ssim"])
    
    
    values = [float(x) for x in list(dict_val.values())]
    


    return file_name, values[0]




def load_files(input_path, ext):
    
    
    patterns = [os.path.join(input_path, f"*.{p}") for p in ext]
    
    files = [f for fs in [glob.glob(pattern) for pattern in patterns] for f in fs]
    
    return sorted(files)


        

if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-i", "--input_path", required = True, help = "path to csv file")
    ap.add_argument("-o", "--output_path", required = False, help = "path to result file")
    ap.add_argument("-ft_csv", "--filetype_csv", required = False,  default = "csv", help = "filetype csv")
    ap.add_argument("-ft_img", "--filetype_img", required = False,  default = "png", help = "filetype image")
    args = vars(ap.parse_args())
    
    
    # setting path to model file
    # input path
    input_path = args["input_path"]
    
    
    
    ext_csv = args['filetype_csv'].split(',') if 'filetype_csv' in args else []
    
    ext_img = args['filetype_img'].split(',') if 'filetype_img' in args else []

    
    #patterns = [os.path.join(input_path, f"*.{p}") for p in ext]
    
    #files = [f for fs in [glob.glob(pattern) for pattern in patterns] for f in fs]
    

    csv_files = load_files(input_path, ext_csv)
    
    img_files = load_files(input_path, ext_img)
    
    print("Input csv files: {}\n".format(csv_files))
    
    print("Input image files: {}\n".format(img_files))
    

    
    # result path
    result_path = args["output_path"] if args["output_path"] is not None else input_path

    result_path = os.path.join(result_path, '')

    # printout result path
    print("Output path: {}\n".format(result_path))
    

    image_mask = cv2.imread(img_files[0])
    
    image_rgb = cv2.imread(img_files[0])
    

    # load csv data
    df = pd.read_csv(csv_files[0], header = None)
    
    print(df, len(df))
    
    corners_bbbox = []
    
    for row_id in range(len(df)):
    
        row_data = df.loc[row_id, :].values.tolist()
        
        print(row_data)
        
        # Coordinates of the 4 corners
        corners = [(row_data[0], row_data[1]), (row_data[2], row_data[3]), (row_data[4], row_data[5]), (row_data[6], row_data[7])]
        
        # Find extreme coordinates to define rectangle
        x_coords = [x for x, y in corners]
        y_coords = [y for x, y in corners]
        
        
        top_left = (min(x_coords), min(y_coords))
        bottom_right = (max(x_coords), max(y_coords))
        
        
        offset = 20
        '''
        cropped_mask = image_mask[min(y_coords)-offset:max(y_coords)+ offset, min(x_coords)-offset:max(x_coords)+offset]
        
        formatted_id = "{:02d}".format(row_id)
        
        save_path = result_path + "/" + str(formatted_id) + "/"
        
        mkdir(save_path)

        result_file = (save_path + str(formatted_id) + '_mask.png')
        
        cv2.imwrite(result_file, cropped_mask)
        '''
        
        cropped_rgb = image_rgb[min(y_coords)-offset:max(y_coords)+ offset, min(x_coords)-offset:max(x_coords)+offset]
        
        formatted_id = "{:02d}".format(row_id)
        
        #save_path = result_path + "/" + str(formatted_id) + "/"
        
        #mkdir(result_path)

        result_file = (result_path + str(formatted_id) + '.png')
        
        cv2.imwrite(result_file, cropped_rgb)
        
        
        # Define the color of the points (BGR format)
        color = (0, 0, 255)  # Red

        # Define the radius of the points
        radius = 2
        
        # Draw the points
        for point in corners:
            result_img = cv2.circle(image_rgb, point, radius, color, -1)
                
        corners_bbbox.append(corners)
        
        print(x_coords)
        print(y_coords)
        
        print(top_left)
        
        print(bottom_right)
        

        # Draw rectangle
        color = (255, 0, 0)
        thickness = 2
        result_img = cv2.rectangle(image_rgb, top_left, bottom_right, color, thickness)
    
    
    
    


    result_file = (result_path + 'bbox.png')
    cv2.imwrite(result_file, result_img)
    

    
    #trait_img = cv2.line(img[3], points[-1], points[-2], (0, 0, 10), 2)
    
    

        
    '''
    #accquire image file list
    filetype = '*.' + ext
    org_image_path = org_file_path + filetype
    pred_image_path = pred_file_path + filetype
    
    #accquire image file list
    org_imgList = sorted(glob.glob(org_image_path))
    pred_imgList = sorted(glob.glob(pred_image_path))
    

    
    #global save_path
    
    n_images = len(org_imgList)
    
    result_list = []

    file_name_list = []
    
    
        
    #loop execute
    
    for i, (image_ori, image_pred) in enumerate(zip(org_imgList, pred_imgList)):
        
        
        (file_name, values) = similarity_evaluation(image_ori, image_pred)
        
        
        result_list.append([file_name, values])

    
    #print(result_list)
    
    #print(type(result_list[0]))

    
    #########################################################################
    #trait_file = (os.path.dirname(os.path.abspath(file_path)) + '/' + 'trait.xlsx')
    
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
   

    
    trait_file = (org_file_path + '/ssim_value.xlsx')

    
    
    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
        

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        #sheet_leaf = wb.create_sheet()

        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'ssim'

        
    for row in result_list:
        sheet.append(row)
   

    #save the csv file
    wb.save(trait_file)
    '''
    
